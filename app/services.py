from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
import os
import re
import shutil
import subprocess
import sys
import tempfile
import textwrap
from typing import Iterable
from uuid import uuid4
import zipfile

from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageFont, ImageOps
from sqlalchemy.orm import Session

from .config import FILES_DIR
from .models import Buyer, Invoice, InvoiceItem

SPECIAL_ELECTRONIC_VAT_RATE = 0.13
STANDARD_ELECTRONIC_VAT_RATE = 0.01
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".tif", ".tiff"}


def infer_invoice_number_from_filename(filename: str | None) -> str:
    if not filename:
        return ""

    stem = Path(filename).stem
    if not stem:
        return ""

    # Prefer long pure-digit segments (common in e-invoice filenames).
    digit_groups = re.findall(r"\d{8,20}", stem)
    if digit_groups:
        return digit_groups[-1]

    # Fallback to alpha-numeric invoice-like token.
    token_groups = re.findall(r"[A-Za-z0-9-]{6,}", stem)
    if token_groups:
        return token_groups[-1]

    return ""


def get_invoice_tax_rate(invoice_type: str) -> float:
    if invoice_type == "增值税发票" or "专用发票" in invoice_type or "电专" in invoice_type:
        return SPECIAL_ELECTRONIC_VAT_RATE
    return STANDARD_ELECTRONIC_VAT_RATE


def resolve_line_item_amounts(
    *,
    quantity: float,
    unit_price_input: str | None,
    amount_with_tax_input: float,
) -> tuple[float, float]:
    quantity_val = float(quantity or 0.0)
    amount_val = round(float(amount_with_tax_input or 0.0), 2)
    unit_text = (unit_price_input or "").strip()

    if unit_text:
        unit_price_val = round(float(unit_text), 6)
        amount_val = round(unit_price_val * quantity_val, 2)
        return unit_price_val, amount_val

    if quantity_val > 0:
        unit_price_val = round(amount_val / quantity_val, 6)
        return unit_price_val, amount_val

    return 0.0, amount_val


def archive_invoice_file(upload_file) -> tuple[str, str] | tuple[None, None]:
    if not upload_file or not upload_file.filename:
        return None, None

    return archive_invoice_file_bytes(upload_file.filename, upload_file.file.read())


def archive_invoice_file_bytes(
    filename: str | None,
    content: bytes | None,
    *,
    bucket_name: str = "",
) -> tuple[str, str] | tuple[None, None]:
    if not filename or content is None:
        return None, None

    month_dir = FILES_DIR / datetime.now().strftime("%Y-%m")
    if bucket_name:
        month_dir = FILES_DIR / bucket_name / datetime.now().strftime("%Y-%m")
    month_dir.mkdir(parents=True, exist_ok=True)

    ext = Path(filename).suffix
    safe_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:8]}{ext}"
    stored_path = month_dir / safe_name

    with stored_path.open("wb") as out:
        out.write(content)

    return filename, str(stored_path)


def create_invoice_with_items(
    db: Session,
    *,
    invoice_type: str,
    invoice_code: str,
    invoice_number: str,
    invoice_date,
    order_number: str,
    order_date,
    tax_rate: float | None,
    seller_id: int,
    buyer_id: int,
    status: str,
    notes: str,
    file_original_name: str | None,
    file_stored_path: str | None,
    item_names: Iterable[str],
    item_specs: Iterable[str],
    item_unit_prices: Iterable[str],
    item_quantities: Iterable[float],
    item_amounts_with_tax: Iterable[float],
) -> Invoice:
    items = []
    amount_total = 0.0
    tax_total = 0.0
    tax_rate_val = tax_rate if tax_rate is not None else get_invoice_tax_rate(invoice_type)

    for name, spec, unit_price_input, qty, total_with_tax_input in zip(
        item_names,
        item_specs,
        item_unit_prices,
        item_quantities,
        item_amounts_with_tax,
    ):
        qty_val = float(qty)
        unit_price_with_tax, total = resolve_line_item_amounts(
            quantity=qty_val,
            unit_price_input=unit_price_input,
            amount_with_tax_input=total_with_tax_input,
        )
        amount = round(total / (1 + tax_rate_val), 2)
        tax_amount = round(total - amount, 2)

        amount_total += amount
        tax_total += tax_amount

        items.append(
            InvoiceItem(
                product_name=name,
                spec_model=spec,
                tax_code="",
                quantity=qty_val,
                unit_price=unit_price_with_tax,
                amount=amount,
                tax_rate=tax_rate_val,
                tax_amount=tax_amount,
                total_with_tax=total,
            )
        )

    invoice = Invoice(
        invoice_type=invoice_type,
        invoice_code=invoice_code,
        invoice_number=invoice_number,
        invoice_date=invoice_date,
        order_number=order_number,
        order_date=order_date,
        seller_id=seller_id,
        buyer_id=buyer_id,
        amount_without_tax=round(amount_total, 2),
        tax_amount=round(tax_total, 2),
        amount_with_tax=round(amount_total + tax_total, 2),
        status=status,
        notes=notes,
        file_original_name=file_original_name,
        file_stored_path=file_stored_path,
        items=items,
    )
    db.add(invoice)
    db.commit()
    db.refresh(invoice)
    return invoice


def export_invoices_xlsx(invoices: list[Invoice], export_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "发票台账"

    ws.append(
        [
            "发票类型",
            "发票号码",
            "开票日期",
            "订单号码",
            "订单日期",
            "销售方",
            "购买方",
            "名称",
            "规格",
            "数量",
            "含税金额",
            "不含税金额",
            "税额",
            "价税合计",
            "状态",
            "原文件名",
            "归档路径",
        ]
    )

    for inv in invoices:
        item_names = "\n".join((item.product_name or "") for item in inv.items)
        item_specs = "\n".join((item.spec_model or "") for item in inv.items)
        item_quantities = "\n".join(str(item.quantity or "") for item in inv.items)
        item_amounts_with_tax = "\n".join(str(item.total_with_tax or "") for item in inv.items)

        ws.append(
            [
                inv.invoice_type,
                inv.invoice_number,
                inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else "",
                inv.order_number or "",
                inv.order_date.strftime("%Y-%m-%d") if inv.order_date else "",
                inv.seller.name if inv.seller else "",
                inv.buyer.name if inv.buyer else "",
                item_names,
                item_specs,
                item_quantities,
                item_amounts_with_tax,
                inv.amount_without_tax,
                inv.tax_amount,
                inv.amount_with_tax,
                inv.status,
                inv.file_original_name or "",
                inv.file_stored_path or "",
            ]
        )

    wb.save(export_path)
    return export_path


def _safe_export_name(value: str | None, fallback: str) -> str:
    text = (value or "").strip()
    if not text:
        return fallback
    sanitized = re.sub(r'[\\/:*?"<>|]+', "_", text)
    sanitized = re.sub(r"\s+", "_", sanitized).strip("._")
    return sanitized or fallback


def _load_font(size: int, *, bold: bool = False) -> ImageFont.ImageFont:
    candidates = []
    if sys.platform == "darwin":
        candidates.extend(
            [
                "/System/Library/Fonts/PingFang.ttc",
                "/System/Library/Fonts/STHeiti Light.ttc",
                "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
            ]
        )
    candidates.extend(
        [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
            "C:/Windows/Fonts/msyh.ttc",
            "C:/Windows/Fonts/simhei.ttf",
        ]
    )

    for path in candidates:
        if not os.path.exists(path):
            continue
        try:
            return ImageFont.truetype(path, size=size)
        except OSError:
            continue
    return ImageFont.load_default()


def _measure_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> int:
    if not text:
        return 0
    bbox = draw.textbbox((0, 0), text, font=font)
    return bbox[2] - bbox[0]


def _wrap_lines(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    paragraphs = (text or "").splitlines() or [""]
    wrapped: list[str] = []
    for paragraph in paragraphs:
        current = ""
        for char in paragraph or " ":
            trial = f"{current}{char}" if current else char
            if current and _measure_text(draw, trial, font) > max_width:
                wrapped.append(current)
                current = char
            else:
                current = trial
        wrapped.append(current or "")
    return wrapped or [""]


def _new_canvas(width: int, height: int) -> tuple[Image.Image, ImageDraw.ImageDraw]:
    image = Image.new("RGB", (width, height), "#ffffff")
    return image, ImageDraw.Draw(image)


def _image_to_png_bytes(image: Image.Image) -> bytes:
    buffer = BytesIO()
    image.save(buffer, format="PNG")
    return buffer.getvalue()


def render_text_image(title: str, body: str, *, subtitle: str = "") -> bytes:
    width = 1440
    padding = 52
    title_font = _load_font(42, bold=True)
    subtitle_font = _load_font(24)
    body_font = _load_font(26)
    _, draw = _new_canvas(width, 400)

    body_lines = _wrap_lines(draw, body or "-", body_font, width - padding * 2)
    content_height = padding + 68
    if subtitle:
        content_height += 40
    content_height += len(body_lines) * 38 + padding

    image, draw = _new_canvas(width, max(400, content_height))
    draw.rounded_rectangle((24, 24, width - 24, image.height - 24), radius=28, fill="#f8fbff", outline="#d8e2f0", width=2)
    draw.text((padding, padding), title, fill="#0f172a", font=title_font)
    y = padding + 68
    if subtitle:
      draw.text((padding, y), subtitle, fill="#64748b", font=subtitle_font)
      y += 40
    for line in body_lines:
        draw.text((padding, y), line, fill="#1e293b", font=body_font)
        y += 38
    return _image_to_png_bytes(image)


def render_invoice_export_image(invoice: Invoice) -> bytes:
    width = 1800
    padding = 44
    title_font = _load_font(44, bold=True)
    section_font = _load_font(24)
    meta_font = _load_font(26)
    header_font = _load_font(24, bold=True)
    cell_font = _load_font(24)
    total_font = _load_font(30, bold=True)
    preview_image, preview_draw = _new_canvas(width, 800)

    meta_lines = [
        f"发票号：{invoice.invoice_number or '-'}    类型：{invoice.invoice_type or '-'}    状态：{invoice.status or '-'}",
        f"开票日期：{invoice.invoice_date.strftime('%Y-%m-%d') if invoice.invoice_date else '-'}    订单号：{invoice.order_number or '-'}",
        f"销售方：{invoice.seller.name if invoice.seller else '-'}",
        f"购买方：{invoice.buyer.name if invoice.buyer else '-'}",
    ]

    columns = [
        ("名称", 300),
        ("规格", 220),
        ("税收编码", 180),
        ("数量", 130),
        ("单价", 150),
        ("未税金额", 180),
        ("税率", 120),
        ("税额", 150),
        ("总金额", 150),
    ]
    wrapped_rows: list[list[list[str]]] = []
    for item in invoice.items:
        row = [
            _wrap_lines(preview_draw, item.product_name or "-", cell_font, columns[0][1] - 20),
            _wrap_lines(preview_draw, item.spec_model or "-", cell_font, columns[1][1] - 20),
            _wrap_lines(preview_draw, item.tax_code or "-", cell_font, columns[2][1] - 20),
            [f"{item.quantity or 0:g}"],
            [f"{float(item.unit_price or 0):g}"],
            [f"{float(item.amount or 0):.2f}"],
            [f"{float(item.tax_rate or 0):g}"],
            [f"{float(item.tax_amount or 0):.2f}"],
            [f"{float(item.total_with_tax or 0):.2f}"],
        ]
        wrapped_rows.append(row)

    table_y = 260
    row_heights = []
    for row in wrapped_rows:
        max_lines = max(len(cell_lines) for cell_lines in row)
        row_heights.append(max(60, 22 + max_lines * 34))

    image_height = table_y + 66 + sum(row_heights) + 150
    image, draw = _new_canvas(width, image_height)
    draw.rounded_rectangle((18, 18, width - 18, image_height - 18), radius=28, fill="#ffffff", outline="#d6e0ec", width=2)
    draw.text((padding, padding), "发票导出", fill="#0f172a", font=title_font)
    draw.text((padding, padding + 62), f"导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", fill="#64748b", font=section_font)

    y = padding + 116
    for line in meta_lines:
        draw.text((padding, y), line, fill="#334155", font=meta_font)
        y += 38

    x = padding
    draw.rounded_rectangle((padding, table_y, width - padding, table_y + 58), radius=18, fill="#edf3ff")
    for title, col_width in columns:
        draw.text((x + 14, table_y + 16), title, fill="#475569", font=header_font)
        x += col_width

    current_y = table_y + 58
    for row, row_height in zip(wrapped_rows, row_heights):
        draw.rectangle((padding, current_y, width - padding, current_y + row_height), fill="#ffffff", outline="#e2e8f0")
        x = padding
        for (title, col_width), cell_lines in zip(columns, row):
            text_y = current_y + 14
            for line in cell_lines:
                draw.text((x + 14, text_y), line, fill="#0f172a", font=cell_font)
                text_y += 32
            x += col_width
        current_y += row_height

    subtotal_y = current_y + 18
    draw.text((padding, subtotal_y), f"小计（不含税）：{float(invoice.amount_without_tax or 0):.2f}", fill="#0f172a", font=total_font)
    draw.text((padding + 520, subtotal_y), f"税额：{float(invoice.tax_amount or 0):.2f}", fill="#0f172a", font=total_font)
    draw.text((padding + 900, subtotal_y), f"价税合计：{float(invoice.amount_with_tax or 0):.2f}", fill="#0f172a", font=total_font)

    if not wrapped_rows:
        draw.text((padding, table_y + 88), "无商品明细", fill="#64748b", font=cell_font)

    return _image_to_png_bytes(image)


def _normalize_image_file(source_path: Path) -> bytes:
    with Image.open(source_path) as image:
        normalized = ImageOps.exif_transpose(image)
        if normalized.mode not in {"RGB", "RGBA"}:
            normalized = normalized.convert("RGBA" if "A" in normalized.getbands() else "RGB")
        return _image_to_png_bytes(normalized)


def _quicklook_preview_bytes(source_path: Path) -> list[bytes]:
    if sys.platform != "darwin":
        return []
    with tempfile.TemporaryDirectory() as temp_dir:
        result = subprocess.run(
            ["/usr/bin/qlmanage", "-t", "-s", "2400", "-o", temp_dir, str(source_path)],
            capture_output=True,
            text=True,
        )
        if result.returncode != 0:
            return []
        preview_dir = Path(temp_dir)
        return [path.read_bytes() for path in sorted(preview_dir.glob("*.png"))]


def export_file_as_images(source_path: Path, *, title: str, subtitle: str = "") -> list[bytes]:
    if not source_path.exists():
        return [render_text_image(title, "原始文件不存在，无法导出。", subtitle=subtitle)]

    suffix = source_path.suffix.lower()
    if suffix in IMAGE_EXTENSIONS:
        return [_normalize_image_file(source_path)]

    if suffix == ".txt":
        try:
            text = source_path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text = source_path.read_text(encoding="utf-8", errors="ignore")
        return [render_text_image(title, text or "(空文本)", subtitle=subtitle)]

    preview_images = _quicklook_preview_bytes(source_path)
    if preview_images:
        return preview_images

    return [
        render_text_image(
            title,
            f"当前文件暂不支持直接转图片导出。\n文件名：{source_path.name}",
            subtitle=subtitle,
        )
    ]


def export_invoice_image_bundle(invoices: list[Invoice], export_path: Path) -> Path:
    with zipfile.ZipFile(export_path, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        for index, invoice in enumerate(invoices, start=1):
            folder = f"{index:03d}_{_safe_export_name(invoice.invoice_number or str(invoice.id), f'invoice_{invoice.id}') }"
            archive.writestr(f"{folder}/发票明细.png", render_invoice_export_image(invoice))

            if invoice.file_stored_path:
                invoice_source = Path(invoice.file_stored_path)
                invoice_images = export_file_as_images(
                    invoice_source,
                    title="发票原件导出",
                    subtitle=invoice.file_original_name or invoice_source.name,
                )
                for image_index, image_bytes in enumerate(invoice_images, start=1):
                    archive.writestr(f"{folder}/发票原件_{image_index:02d}.png", image_bytes)

            if invoice.trade_voucher_stored_path:
                voucher_source = Path(invoice.trade_voucher_stored_path)
                voucher_images = export_file_as_images(
                    voucher_source,
                    title="凭证导出",
                    subtitle=invoice.trade_voucher_original_name or voucher_source.name,
                )
                for image_index, image_bytes in enumerate(voucher_images, start=1):
                    archive.writestr(f"{folder}/凭证_{image_index:02d}.png", image_bytes)
            elif (invoice.trade_voucher_text or "").strip():
                archive.writestr(
                    f"{folder}/凭证说明.png",
                    render_text_image("凭证说明", invoice.trade_voucher_text or ""),
                )

            manifest = [
                f"发票号: {invoice.invoice_number or '-'}",
                f"订单号: {invoice.order_number or '-'}",
                f"购买方: {invoice.buyer.name if invoice.buyer else '-'}",
                f"销售方: {invoice.seller.name if invoice.seller else '-'}",
                f"商品条数: {len(invoice.items)}",
            ]
            for row_index, item in enumerate(invoice.items, start=1):
                manifest.append(
                    f"{row_index}. {(item.product_name or '-')} | {(item.spec_model or '-')} | 数量 {(item.quantity or 0):g} | 总金额 {float(item.total_with_tax or 0):.2f}"
                )
            archive.writestr(f"{folder}/清单.txt", "\n".join(manifest))

    return export_path


def export_customer_profile_xlsx(
    *,
    buyer: Buyer,
    invoices: list[Invoice],
    product_category_map: dict[str, str],
    export_path: Path,
) -> Path:
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "客户画像"

    total_amount = round(sum(float(inv.amount_with_tax or 0.0) for inv in invoices), 2)
    total_count = len(invoices)
    total_quantity = 0.0
    first_purchase_date = ""
    last_purchase_date = ""
    if invoices:
        sorted_invoices = sorted(invoices, key=lambda inv: (inv.invoice_date or datetime.today().date(), inv.id))
        first_purchase_date = sorted_invoices[0].invoice_date.strftime("%Y-%m-%d") if sorted_invoices[0].invoice_date else ""
        last_purchase_date = sorted_invoices[-1].invoice_date.strftime("%Y-%m-%d") if sorted_invoices[-1].invoice_date else ""
    avg_amount = round((total_amount / total_count), 2) if total_count else 0.0

    product_stats: dict[str, dict[str, float]] = {}
    category_stats: dict[str, dict[str, float]] = {}
    recent_details_rows = []
    product_frequency: dict[str, int] = {}

    for inv in sorted(invoices, key=lambda inv: (inv.invoice_date or datetime.today().date(), inv.id), reverse=True):
        invoice_products = set()
        order_qty = 0.0
        detail_names = []
        for item in inv.items:
            product_name = (item.product_name or "未命名商品").strip() or "未命名商品"
            category_name = product_category_map.get(product_name, "未分组")
            qty = float(item.quantity or 0.0)
            amount = float(item.total_with_tax or 0.0)
            total_quantity += qty
            order_qty += qty
            detail_names.append(product_name)
            invoice_products.add(product_name)

            product_stats.setdefault(product_name, {"amount": 0.0, "quantity": 0.0, "frequency": 0})
            product_stats[product_name]["amount"] += amount
            product_stats[product_name]["quantity"] += qty

            category_stats.setdefault(category_name, {"amount": 0.0, "quantity": 0.0})
            category_stats[category_name]["amount"] += amount
            category_stats[category_name]["quantity"] += qty

        for product_name in invoice_products:
            product_frequency[product_name] = product_frequency.get(product_name, 0) + 1

        recent_details_rows.append(
            [
                inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else "",
                inv.invoice_number,
                "、".join(detail_names[:3]) + (" 等" if len(detail_names) > 3 else ""),
                round(order_qty, 2),
                round(float(inv.amount_with_tax or 0.0), 2),
                inv.status or "",
            ]
        )

    summary_ws.append(["客户名称", buyer.name])
    summary_ws.append(["累计采购金额", total_amount])
    summary_ws.append(["累计采购数量", round(total_quantity, 2)])
    summary_ws.append(["采购频次", total_count])
    summary_ws.append(["平均客单价", avg_amount])
    summary_ws.append(["首次采购时间", first_purchase_date])
    summary_ws.append(["最近一次采购时间", last_purchase_date])

    product_ws = wb.create_sheet("商品排行")
    product_ws.append(["商品", "累计金额", "累计数量", "购买频次"])
    for name, stats in sorted(product_stats.items(), key=lambda row: row[1]["amount"], reverse=True):
        product_ws.append([
            name,
            round(stats["amount"], 2),
            round(stats["quantity"], 2),
            product_frequency.get(name, 0),
        ])

    category_ws = wb.create_sheet("品类分布")
    category_ws.append(["品类", "累计金额", "累计数量"])
    for name, stats in sorted(category_stats.items(), key=lambda row: row[1]["amount"], reverse=True):
        category_ws.append([name, round(stats["amount"], 2), round(stats["quantity"], 2)])

    detail_ws = wb.create_sheet("最近采购明细")
    detail_ws.append(["日期", "发票号", "商品摘要", "数量", "金额", "状态"])
    for row in recent_details_rows[:20]:
        detail_ws.append(row)

    wb.save(export_path)
    return export_path
